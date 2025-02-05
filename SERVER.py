from flask import Flask, request
from linebot.models import TemplateSendMessage,ButtonsTemplate,MessageAction,URIAction,TextSendMessage,ImageSendMessage,CarouselTemplate,CarouselColumn
import json
from ping3 import ping 
import os
from linebot import LineBotApi, WebhookHandler
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
import pandas as pd
import keyboard 
import time
import requests
import re
import random
from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.utils import get_column_letter

model_path = "/Users/chienchihhsiang/.lmstudio/models/shenzhi-wang/Llama3-8B-Chinese-Chat-GGUF-8bit/Llama3-8B-Chinese-Chat-q8_0-v2_1.gguf"
LLAMA_API_URL = '' 
Material = False
in_process = False
in_process2 = False
pass_process = False
sec_in_process = False
sec_in_process2 = False
unlock_repaire_station = False
image_process = False
unlock_repaire_station2 = False
QRcode_check = False
return_msg = False
internet_process = False
class_process = False
global error_code
error_code = None
line_bot_api = LineBotApi('')
handler = WebhookHandler('')


def check_or_create_excel(file_path="E:/MES自動化/line_bot_log.xlsx"):
    if os.path.exists(file_path):
        print(f"檔案 '{file_path}' 已存在，正在讀取...")
        workbook = load_workbook(file_path)
        sheet = workbook.active
        if sheet.max_row > 1 and sheet.max_column > 1:
            print("第一列和第一行的數據如下：")
            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    print(cell.value, end="\t")
            print()
        else:
            print("檔案存在，但內容是空的。")
    else:
        print(f"檔案 '{file_path}' 不存在，正在創建新檔案...")
        workbook = Workbook()
        sheet = workbook.active

        sheet.title = "Log"
        sheet.append(["Timestamp", "ID", "Event", "Department", "Class", "Serial Number"])
        print("已創建標題行。")

        os.makedirs(os.path.dirname(file_path), exist_ok=True)  # 確保路徑存在
        workbook.save(file_path)
        print(f"新檔案 '{file_path}' 已創建並儲存。")

    print("操作完成！")

AT = "U1d2fae8c03fbc0819fe079da5250d9bc"

def bug_c_automation(work_order,input_QR):
    global actions2
    options = Options()
    options.add_argument('--ignore-certificate-errors')  
    options.add_argument('--disable-web-security')      
    input_acc = "14574"
    driver = webdriver.Chrome(options=options)
    driver.set_window_size(1920, 1080)
    driver.maximize_window()
    actions2 = ActionChains(driver)
    wait = WebDriverWait(driver, 10)  
    wait2 = WebDriverWait(driver, 300)  
    try:
        driver.get(f"http://cimes.seec.com.tw/{work_order}/CimesDesktop.aspx")
        username_input = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input#UserName"))
        )
        password_input = driver.find_element(By.CSS_SELECTOR, "input#Password")
        login_button = driver.find_element(By.ID, "LoginButton")
        username_input.send_keys(input_acc)
        password_input.send_keys(input_acc)
        login_button.click()
        all_windows = driver.window_handles
        for window in driver.window_handles:
            if window != all_windows:
                driver.switch_to.window(window)
                break
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="QryProg"]')))
        element4.click()
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//span[text()="在製品查詢"]')))
        actions.move_to_element(button).click().perform()
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        keyboard.press_and_release('ctrl+-')
        keyboard.press_and_release('ctrl+-')
        element6 = wait2.until(EC.presence_of_element_located((By.ID, "btnLotQuery")))
        element6.click()
        iframe2 = wait.until(EC.presence_of_element_located((By.XPATH, '//iframe[contains(@src, "LotListFieldSelect.aspx")]')))
        driver.switch_to.frame(iframe2)
        checkbox = wait2.until(EC.presence_of_element_located((By.ID, "tvn0CheckBox")))
        if checkbox.is_selected():
            checkbox.click()
        else:
            pass
        input_element3 = wait.until(EC.presence_of_element_located((By.ID, "gvFilter_ctl03_ttbValue")))
        input_element3.clear()
        input_element3.send_keys(input_QR)
        Query = wait.until(EC.presence_of_element_located((By.ID, "btnQuery")))
        Query.click()
        time.sleep(1)
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        
        if work_order in ["ARG","AMG"]:
            try:
                elementQR = wait.until(EC.presence_of_element_located((By.XPATH, f"//td[text()='{input_QR}']")))
                elementQR.click()
            except:
                error_code = "1"
                return error_code
        elif work_order == "AM2":
            try:
                element7 = wait.until(EC.presence_of_element_located((By.ID, "gvDataViewer_ctl02_hlLot")))
                driver.execute_script("arguments[0].click();", element7)
                actions.move_to_element(element7).click().perform()
                Niframes = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
                driver.switch_to.frame(Niframes[0])
                element8 = wait.until(EC.presence_of_element_located((By.ID, "btnExit")))
                element8.click()

            except:
                error_code = "1"
                return error_code



        time.sleep(1)
        driver.switch_to.default_content()
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()

        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="WipRule_Modify"]')))
        element4.click()    
        button2 = wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(@title, '變更工作站')]")))
        actions.move_to_element(button2).click().perform()
        time.sleep(1)
        driver.switch_to.default_content()
        Giframes = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
        driver.switch_to.frame(Giframes[1])
        select_list = wait.until(EC.presence_of_element_located((By.ID, "_ddlcsOperation")))
        select = Select(select_list)
        select.select_by_index(2)
        time.sleep(0.5)
        select_list2 = wait.until(EC.presence_of_element_located((By.ID, "_ddlcsReason")))
        select2 = Select(select_list2)
        select2.select_by_index(1)
        time.sleep(0.5)
        finish_btnOK = wait.until(EC.presence_of_element_located((By.ID, "btnOK")))
        finish_btnOK.click()
        time.sleep(5)    
    finally:
        driver.quit()


def requests_website_test(work_order,tk):
    url = f"http://cimes.seec.com.tw/{work_order}/CimesDesktop.aspx"
    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            line_bot_api.reply_message(tk,TextSendMessage(f"連線良好，{response.elapsed.total_seconds()} 秒"))
        else:
            line_bot_api.reply_message(tk,TextSendMessage(f"連線不良，{response.elapsed.total_seconds()} 秒"))
    except requests.exceptions.RequestException as e:
        line_bot_api.reply_message(tk,TextSendMessage(f"連線失敗"))





def repair_automation(work_order,input_QR,class_number):
    global actions2
    global error_code
    global error_text
    error_text = None
    error_code = False
    options = Options()
    options.add_argument('--ignore-certificate-errors')  
    options.add_argument('--disable-web-security')      
    input_acc = "14574"
    driver = webdriver.Chrome(options=options)
    driver.set_window_size(1920, 1080)
    driver.maximize_window()
    actions2 = ActionChains(driver)
    wait = WebDriverWait(driver, 10)  
    wait2 = WebDriverWait(driver, 300)  
    try:
        global last_column
        driver.get(f"http://cimes.seec.com.tw/{work_order}/CimesDesktop.aspx")
        username_input = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input#UserName"))
        )
        password_input = driver.find_element(By.CSS_SELECTOR, "input#Password")
        login_button = driver.find_element(By.ID, "LoginButton")
        username_input.send_keys(input_acc)
        password_input.send_keys(input_acc)
        login_button.click()
        all_windows = driver.window_handles
        for window in driver.window_handles:
            if window != all_windows:
                driver.switch_to.window(window)
                break
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="WipRule"]')))
        element4.click()
        button2 = wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(@title, '維修(客)')]")))
        actions.move_to_element(button2).click().perform()
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        keyboard.press_and_release('ctrl+-')
        keyboard.press_and_release('ctrl+-')
        input1 = wait.until(EC.presence_of_element_located((By.ID, "ttbLot")))
        input1.send_keys(input_QR)
        input1.send_keys(Keys.RETURN)
        time.sleep(1)
        try:
            error_element = wait2.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".exceptionTitle")))
            error_text = error_element.text    
            close_button = wait2.until(EC.element_to_be_clickable(
                (By.XPATH, "//td//input[@value='關閉' and @type='button']")
            ))
            driver.execute_script("arguments[0].click();", close_button)
            error_code = True
            return
        except:
            pass
        select_OTHER = wait.until(EC.presence_of_element_located((By.ID, "ddlRepairFunction")))
        selectlist3 = Select(select_OTHER)
        selectlist3.select_by_visible_text("[OTHER]")
        select_list = wait.until(EC.presence_of_element_located((By.ID, "ddlDutyUnit")))
        enable_input2 = wait.until(EC.presence_of_element_located((By.NAME, "ttbDefectOperation")))
        last_column = enable_input2.get_attribute('value')
        select_list = wait.until(EC.presence_of_element_located((By.ID, "ddlDutyUnit")))
        select = Select(select_list)
        select.select_by_visible_text(class_number)
        select_list2 = wait.until(EC.presence_of_element_located((By.ID, "ddlReturnOperation")))
        select = Select(select_list2)
        select.select_by_visible_text(last_column)
        input2 = wait.until(EC.presence_of_element_located((By.ID, "ttbRepairTime")))
        input2.send_keys("0")
        time.sleep(2)
        finish_btnOK = wait.until(EC.presence_of_element_located((By.ID, "btnOK")))
        finish_btnOK.click()
        time.sleep(1)    
    finally:
        driver.quit()




def work_order_automation(work_order,data_input):
    global wait2,actions2
    global value
    options = Options()
    options.add_argument('--ignore-certificate-errors')  
    options.add_argument('--disable-web-security')      
    input_acc = "14574"
    driver = webdriver.Chrome(options=options)
    driver.set_window_size(1920, 1080)
    driver.maximize_window()
    actions2 = ActionChains(driver)
    wait = WebDriverWait(driver, 10)  
    wait2 = WebDriverWait(driver, 20)  
    try:
        driver.get(f"http://cimes.seec.com.tw/{work_order}/Security/CimesUserLogin.aspx")
        username_input = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input#UserName"))
        )
        password_input = driver.find_element(By.CSS_SELECTOR, "input#Password")
        login_button = driver.find_element(By.ID, "LoginButton")
        username_input.send_keys(input_acc)
        password_input.send_keys(input_acc)
        login_button.click()
        all_windows = driver.window_handles

        for window in driver.window_handles:
            if window != all_windows:
                driver.switch_to.window(window)
                break
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="CustRule"]')))
        element4.click()
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//span[text()="日工單開立"]')))
        actions.move_to_element(button).click().perform()
        time.sleep(3)
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        input_element3 = wait.until(EC.presence_of_element_located((By.ID, "CimesInputBox")))
        keyboard.press_and_release('ctrl+-')
        input_element3.send_keys(data_input)
        input_element3.send_keys(Keys.RETURN)
        time.sleep(2)
        enable_input = wait.until(EC.presence_of_element_located((By.ID, "ttbUnCreateQty")))
        value = enable_input.get_attribute('value')
        print(type(value))
        if value == '0':
            return value
        input_element4 = wait.until(EC.presence_of_element_located((By.ID, "ttbDayWOQty")))
        input_element4.send_keys(value)
        add_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "btnAdd"))
        )                
        add_button.click()
        ok_button = wait.until(
        EC.element_to_be_clickable((By.ID, "btnOK"))
        )             
        actions.move_to_element(ok_button).click().perform()
        time.sleep(3)    
    finally:
        driver.quit()




def broadcast():
    message = "【公告】系統即將進行維護，造成不便請見諒。"
    line_bot_api.broadcast(TextSendMessage(text=message))

def image(user_id,content):
    image_url = f"https://api.qrserver.com/v1/create-qr-code/?size=200x200&data={content}"
    image_message = ImageSendMessage(
    original_content_url=image_url,  
    preview_image_url=image_url     
)
    line_bot_api.push_message(user_id, image_message)


def select_web(tk):
    line_bot_api.reply_message(
            tk,
            [
                TemplateSendMessage(
                    alt_text='部門選擇',
                    template=ButtonsTemplate(
                        thumbnail_image_url='https://i.ibb.co/Qj1b4KL/2.jpg',
                        title='部門選擇',
                        text='請從以下選項選擇一個部門：',
                        actions=[
                            MessageAction(label='ARG', text='ARG'),
                            MessageAction(label='AMG', text='AMG'),
                            MessageAction(label='AM2', text='AM2')
                        ]
                    )
                ),
                TextSendMessage(text="請選擇部門，或輸入[返回]結束動作")
            ]
        )
def select_web2(tk):
    line_bot_api.reply_message(
            tk,
            [
                TemplateSendMessage(
                    alt_text='班別選擇',
                    template=ButtonsTemplate(
                        thumbnail_image_url='https://i.ibb.co/cyBPJHW/2.jpg',
                        title='班別選擇',
                        text='請從以下選項選擇一個班別：',
                        actions=[
                            MessageAction(label='AF1', text='AF1'),
                            MessageAction(label='AF2', text='AF2'),
                            MessageAction(label='AF3', text='AF3')
                        ]
                    )
                ),
                TextSendMessage(text="請選擇班別，或輸入[返回]結束動作")
            ]
        )


def select_web3(tk):
    line_bot_api.reply_message(
            tk,
            [
                TemplateSendMessage(
                    alt_text='班別選擇',
                    template=ButtonsTemplate(
                        thumbnail_image_url='https://i.ibb.co/KFFgHK1/3.jpg',
                        title='班別選擇',
                        text='請從以下選項選擇一個班別：',
                        actions=[
                            MessageAction(label='AC1', text='AC1'),
                            MessageAction(label='AC2', text='AC2'),
                            MessageAction(label='AC3', text='AC3')
                        ]
                    )
                ),
                TextSendMessage(text="請選擇班別，或輸入[返回]結束動作")
            ]
        )


def select_web4(tk):
    line_bot_api.reply_message(
        tk,
        [
            TemplateSendMessage(
                alt_text='班別選擇 1',
                template=ButtonsTemplate(
                    thumbnail_image_url='https://i.ibb.co/z2sW8TF/image.jpg',
                    title='班別選擇 1',
                    text='請從以下選項選擇一個班別：',
                    actions=[
                        MessageAction(label='AR1', text='AR1'),
                        MessageAction(label='AR2', text='AR2'),
                        MessageAction(label='AR3', text='AR3')
                    ]
                )
            ),
            TemplateSendMessage(
                alt_text='班別選擇 2',
                template=ButtonsTemplate(
                    thumbnail_image_url='https://i.ibb.co/z2sW8TF/image.jpg',  
                    title='班別選擇 2',
                    text='請從以下選項選擇其他班別：',
                    actions=[
                        MessageAction(label='AR4', text='AR4'),
                        MessageAction(label='AF1', text='AF1'),
                        MessageAction(label='返回', text='返回')
                    ]
                )
            ),
            TextSendMessage(text="請選擇班別，或輸入[返回]結束動作")
        ]
    )





def send_carousel_message(user_id):
    line_bot_api.push_message(user_id, TemplateSendMessage(
        alt_text='Menu',
        template=CarouselTemplate(
            columns=[
                CarouselColumn(
                    thumbnail_image_url='https://i.ibb.co/Qj1b4KL/2.jpg',
                    title='工單系統',
                    text='以下為工單相關服務',
                    actions=[
                        MessageAction(
                            label='開立工單',
                            text='開立工單'
                        ),
                        MessageAction(
                            label='解除維修站',
                            text='解除維修站'
                        ),
                        MessageAction(
                            label='序號卡Create',
                            text='序號卡Create'
                        )
                    ]
                ),
                CarouselColumn(
                    thumbnail_image_url='https://i.ibb.co/Ngq4L8h/Manufacturing-Execution-System-2.jpg',  # 更新成新的圖片
                    title='其他服務',
                    text='以下為其他相關服務',
                    actions=[
                        URIAction(
                            label='自動排班模型',
                            uri='https://seec-auto.vercel.app/'
                        ),
                        URIAction(
                            label='掃描器',
                            uri='https://line.me/R/nv/QRCodeReader'
                        ),
                        MessageAction(
                            label='QRcode小工具',
                            text='QRcode小工具'
                        )
                    ]
                ),
                CarouselColumn(
                    thumbnail_image_url='https://i.ibb.co/mHd3tpR/2.jpg',  # 網路測試選單圖片
                    title='MES網路測試',
                    text='以下為MES網路測試功能',
                    actions=[
                        MessageAction(
                            label='MES網路測試',
                            text='MES網路測試'
                        ),
                        MessageAction(
                            label='物料QR小工具',
                            text='物料QR小工具'
                        ),
                        MessageAction(
                            label='反應問題',
                            text='反應問題'
                        )
                    ]
                )
            ]
        )
    ))



def monitor_host(host, interval, max_loops):
    response_times = []  # 儲存每次ping的回應時間
    for _ in range(max_loops):
        try:
            response_time = ping(host)
            if response_time is None:
                print(f"{host} is unreachable.")
                response_times.append(None)
            else:
                print(f"{host} response time: {response_time * 1000:.2f} ms")
                response_times.append(response_time * 1000)  
        except Exception as e:
            print(f"An error occurred: {e}")
            response_times.append(None)
        time.sleep(interval)  
    return response_times
    
def check_connection_status(response_times):
    for response_time in response_times:
        if response_time is None or response_time == 0:
            return "無法連線"          
        elif response_time > 500:
            return "無法連線"
        elif response_time > 100:
            return "連線延遲"
    return "連線良好"  


user_ticket_number = None


def replace_number_with_random_iteratively(user_id,text,tk):
    pattern = r";(\d+);[A-Za-z]+"
    edited_text = text  
    generated_numbers = []  
    for _ in range(3):
        match = re.search(pattern, edited_text)
        if match:
            original_number = match.group(1)  
            random_number = random.randint(10000, 99999)
            edited_text = re.sub(rf";{original_number};", f";{random_number};", edited_text, count=1)
            image(user_id = user_id,content = edited_text)
        else:
            line_bot_api.reply_message(tk,TextSendMessage(f"格式可能有誤"))
            break
    return edited_text, generated_numbers

def log_action(file_path="E:/MES自動化/line_bot_log.xlsx", id="", event="",department="", class_number="",Serial_number=""):
    check_or_create_excel(file_path)
    workbook = load_workbook(file_path)
    sheet = workbook.active

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([timestamp, id, event, department, class_number, Serial_number])
    for col_num, column_cells in enumerate(sheet.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        sheet.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    workbook.save(file_path)
    print(f"檔案 '{file_path}' 已更新並儲存。")
check_or_create_excel()

app = Flask(__name__)

@app.route("/", methods=['POST'])
def linebot():

    global class_process
    global class_number
    global work_order
    global user_ticket_number
    global in_process
    global in_process2
    global pass_process
    global sec_in_process
    global sec_in_process2
    global user_ticket_number2
    global unlock_repaire_station
    global unlock_repaire_station_memory
    global unlock_repaire_station2
    global QRcode_check
    global QRcode_content
    global return_msg
    global internet_process
    global Material
    global title_value
    global error_code
    global value
    global image_process
    body = request.get_data(as_text=True)                    
    try:
        json_data = json.loads(body)                         
        access_token = ''
        secret = ''
        line_bot_api = LineBotApi(access_token)              
        handler = WebhookHandler(secret)                    
        signature = request.headers['X-Line-Signature']      
        handler.handle(body, signature)                      
        tk = json_data['events'][0]['replyToken']          
        type = json_data['events'][0]['message']['type']  
        user_id = None   
        user_id = json_data['events'][0]['source']['userId']
        if type == "text" :
            msg = json_data['events'][0]['message']['text']  
            if msg == "開立工單":
                event = msg
                select_web(tk)
                value = None
                title_value = None
                error_code = None
                user_ticket_number = None
                work_order = None
                in_process = True 
            elif in_process:
                if msg == "返回":
                    in_process = False
                else:
                    line_bot_api.push_message(user_id, TextSendMessage("請輸入工單號碼，或輸入[返回]結束動作"))                                            
                    work_order = msg        
                    sec_in_process = True
                    in_process = False 
            elif sec_in_process:
                if msg == "返回":
                    sec_in_process = False
                else:                       
                    user_ticket_number = msg
                    line_bot_api.reply_message(tk, TextSendMessage(f"工單號碼為：{user_ticket_number}，處理中，請稍候"))                
                    try:
                        work_order_automation(work_order = work_order,data_input = user_ticket_number)
                        if value == '0':
                            image(user_id=user_id,content=f"{user_ticket_number}-001")
                            line_bot_api.push_message(user_id, TextSendMessage("工單在月初已開立"))
                        else:    
                            image(user_id=user_id,content=f"{user_ticket_number}-001")
                            line_bot_api.push_message(user_id, TextSendMessage("🎉🎉🎊工單開立完成🎊🎉🎉"))
                            log_action(id=user_id, event="開立工單",department=work_order,Serial_number=value)
                    except:
                        line_bot_api.push_message(user_id, TextSendMessage("查無此工單"))
                        pass
                    sec_in_process = False

            elif msg == "序號卡Create":
                event = msg
                user_ticket_number2 = None
                work_order = None
                select_web(tk)
                in_process2 = True
            elif in_process2:
                if msg == "返回":
                    in_process2 = False
                else:
                    work_order = msg
                    line_bot_api.reply_message(tk, TextSendMessage("請輸入序號，或輸入[返回]結束動作"))
                    in_process2 = False
                    sec_in_process2 = True
            elif sec_in_process2:
                if msg == "返回":
                    sec_in_process2 = False
                else:
                    user_ticket_number2 = msg
                    line_bot_api.reply_message(tk, TextSendMessage(f"序號：{user_ticket_number2}，處理中，請稍候"))
                    try:
                        bug_c_automation(work_order=work_order,input_QR=user_ticket_number2)
                        line_bot_api.push_message(user_id, TextSendMessage("🎉🎊解除卡Create完成🎊🎉請從第二站開始作業，若本來在第二站請從第三站開始作業"))
                        log_action(id=user_id, event="序號卡Create",department=work_order, Serial_number=user_ticket_number2)
                    except:
                        line_bot_api.push_message(user_id, TextSendMessage("系統發生異常"))
                        pass
                    sec_in_process2 = False

            elif msg == "選單":
                send_carousel_message(user_id)

            elif msg == "MES網路測試":
                event = msg
                work_order = None
                select_web(tk)
                internet_process = True
            elif internet_process:
                if msg == "返回":
                    internet_process = False
                else:
                    work_order = msg
                    requests_website_test(work_order = work_order,tk=tk)      
                    internet_process = False            
                    log_action(id=user_id, event="MES網路測試",department=work_order)


            elif msg == "解除維修站":
                event = msg
                select_web(tk)
                work_order = None
                QRcode_content = None
                class_number = None
                unlock_repaire_station_memory = None
                unlock_repaire_station = True
            elif unlock_repaire_station:
                if msg == "返回":
                    unlock_repaire_station = False
                elif msg == "ARG":
                    work_order = msg
                    select_web4(tk)
                    unlock_repaire_station = False
                    unlock_repaire_station2 = True
                elif msg == "AMG":
                    work_order = msg
                    select_web3(tk)
                    unlock_repaire_station = False
                    unlock_repaire_station2 = True
                elif msg == "AM2":
                    work_order = msg
                    select_web2(tk)
                    unlock_repaire_station = False
                    unlock_repaire_station2 = True
                else:    
                    select_web(tk)
            elif unlock_repaire_station2:
                if msg == "返回":
                    unlock_repaire_station2 = False
                elif work_order == "ARG" and (msg == "AR1" or msg == "AR2" or msg == "AR3" or msg == "AR4" or msg == "AF1"):
                    class_number = msg                  
                    line_bot_api.reply_message(tk, TextSendMessage("請輸入序號，或輸入[返回]結束動作"))
                    unlock_repaire_station2 = False
                    class_process = True
                elif work_order == "AMG" and (msg == "AC1" or msg == "AC2" or msg == "AC3"):
                    class_number = msg
                    line_bot_api.reply_message(tk, TextSendMessage("請輸入序號，或輸入[返回]結束動作"))
                    unlock_repaire_station2 = False
                    class_process = True
                elif work_order == "AM2" and (msg == "AF1" or msg == "AF2" or msg == "AF3"):
                    line_bot_api.reply_message(tk, TextSendMessage("請輸入序號，或輸入[返回]結束動作"))
                    unlock_repaire_station2 = False
                    class_number = msg
                    class_process = True
                else:
                    if work_order == "ARG":                      
                        select_web4(tk)
                    elif work_order == "AMG":
                        select_web3(tk)
                    elif work_order == "AM2":
                        select_web2(tk)
                    else:
                        pass
            elif class_process:
                if msg == "返回":
                    class_process = False
                else:
                    unlock_repaire_station_memory = msg
                    line_bot_api.reply_message(tk, TextSendMessage(f"{unlock_repaire_station_memory}處理中，請稍後"))
                    try:
                        error_code = None
                        repair_automation(work_order = work_order,input_QR = unlock_repaire_station_memory,class_number=class_number)
                        line_bot_api.push_message(user_id, TextSendMessage(f"🎉🎉🎊解除維修完成🎊🎉🎉，請從「{last_column}」刷讀")) 
                        log_action(id=user_id, event="解除維修站",department=work_order, class_number=class_number,Serial_number=unlock_repaire_station_memory)
                    except:
                        line_bot_api.push_message(user_id, TextSendMessage("系統發生異常")) 
                        pass
                    class_process = False
            elif error_code:
                line_bot_api.push_message(user_id, TextSendMessage(F"{error_text}"))
                error_code = False
                
            elif msg == "公告":
                line_bot_api.reply_message(tk, TextSendMessage("請輸入密碼："))
                pass_process = True
            elif pass_process:
                if msg == "h123300q795":
                    broadcast()
                    pass_process = False
                else:
                    line_bot_api.push_message(user_id, TextSendMessage(f"驗證失敗"))
                    pass_process = False

            elif msg == "QRcode小工具":     
                event = msg
                line_bot_api.reply_message(tk, TextSendMessage("請輸入內容，或輸入[返回]取消動作"))
                QRcode_check = True
            elif QRcode_check:
                if msg == "返回":
                    QRcode_check = False
                else:
                    QRcode_content = msg
                    image(user_id=user_id,content=QRcode_content)
                    QRcode_check = False
                    try:
                        log_action(id=user_id, event="QRcode小工具", Serial_number=QRcode_content)
                        print("成功")
                    except:
                        print("失敗")
            elif msg == "反應問題":
                line_bot_api.reply_message(tk,TextSendMessage('請輸入要反應的事項，結束請輸入「結束」。'))
                return_msg = True
            elif return_msg:
                if msg == '結束':        
                    line_bot_api.reply_message(tk,TextSendMessage('問題已回傳'))
                    return_msg = False
                else :
                    line_bot_api.push_message(AT, TextSendMessage(msg)) 
                    pass
            elif msg == "物料QR小工具":
                event = msg
                line_bot_api.reply_message(tk, TextSendMessage("請輸入材料編號，或輸入[返回]結束動作"))
                Material = True
            elif Material:
                if msg == "返回":
                    Material = False
                else:
                    text = msg
                    replace_number_with_random_iteratively(user_id=user_id,text=text,tk=tk)
                    log_action(id=user_id, event="物料QR小工具", Serial_number=text)
                    Material = False

            elif msg == "圖像辨識文字":
                line_bot_api.reply_message(tk,TextSendMessage('尚未完成'))
                """image_process = True
            elif image_process:
                if msg == "返回":
                    image_process = False
                else:
                    callback()
                    handle_image_message(event)
                    image_process = False"""
            else:
                line_bot_api.reply_message(tk,TextSendMessage('其他需求請洽：製管MES組'))# 回傳訊息
        else:
            line_bot_api.reply_message(tk,TextSendMessage('你傳的不是文字呦～'))# 回傳訊息

        
    except:
        print(body)                                         
    return 'OK'                                              

if __name__ == "__main__":
    app.run()
